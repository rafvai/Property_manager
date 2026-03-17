import os
from datetime import datetime
from pathlib import Path

from reportlab.lib import colors
from reportlab.lib.pagesizes import A4, landscape
from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
from reportlab.lib.units import cm
from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer
from reportlab.lib.enums import TA_CENTER, TA_RIGHT, TA_LEFT
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
from openpyxl.utils import get_column_letter

from config import Config
from styles import COLORE_ERROR


class ExportService:
    """Gestisce l'export di transazioni in PDF e Excel"""

    def __init__(self):
        # FIX: usa Config.EXPORTS_DIR invece di "exports" hardcoded
        self.exports_dir = Config.EXPORTS_DIR or Path('exports').absolute()
        self.exports_dir.mkdir(parents=True, exist_ok=True)

    def export_to_pdf(self, transactions, property_name=None, start_date=None, end_date=None):
        """Esporta transazioni in PDF"""
        timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
        filepath  = self.exports_dir / f"transazioni_{timestamp}.pdf"

        doc = SimpleDocTemplate(
            str(filepath),
            pagesize    =landscape(A4),
            rightMargin =1.5 * cm,
            leftMargin  =1.5 * cm,
            topMargin   =2 * cm,
            bottomMargin=2 * cm
        )

        elements = []
        styles   = getSampleStyleSheet()

        title_style = ParagraphStyle(
            'CustomTitle',
            parent   =styles['Heading1'],
            fontSize =18,
            textColor=colors.HexColor('#1e7be7'),
            spaceAfter=20,
            alignment=TA_CENTER
        )
        elements.append(Paragraph("📊 Report Transazioni", title_style))

        info_style = ParagraphStyle(
            'Info',
            parent   =styles['Normal'],
            fontSize =10,
            textColor=colors.HexColor('#666666'),
            alignment=TA_CENTER
        )
        info_text = f"Generato il: {datetime.now().strftime('%d/%m/%Y %H:%M')}"
        if property_name:
            info_text += f" | Proprietà: {property_name}"
        if start_date and end_date:
            info_text += f" | Periodo: {start_date} - {end_date}"
        elements.append(Paragraph(info_text, info_style))
        elements.append(Spacer(1, 0.5 * cm))

        totale_entrate = sum(t['amount'] for t in transactions if t['type'] == 'Entrata')
        totale_uscite  = sum(t['amount'] for t in transactions if t['type'] == 'Uscita')
        saldo          = totale_entrate - totale_uscite

        summary_data  = [
            ['Totale Entrate', 'Totale Uscite', 'Saldo Netto'],
            [f'€ {totale_entrate:,.2f}', f'€ {totale_uscite:,.2f}', f'€ {saldo:,.2f}']
        ]
        summary_table = Table(summary_data, colWidths=[6 * cm, 6 * cm, 6 * cm])
        summary_table.setStyle(TableStyle([
            ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#34495e')),
            ('TEXTCOLOR',  (0, 0), (-1, 0), colors.whitesmoke),
            ('ALIGN',      (0, 0), (-1, -1), 'CENTER'),
            ('FONTNAME',   (0, 0), (-1, 0), 'Helvetica-Bold'),
            ('FONTSIZE',   (0, 0), (-1, 0), 12),
            ('BOTTOMPADDING', (0, 0), (-1, 0), 12),
            ('BACKGROUND', (0, 1), (0, 1), colors.HexColor('#d4edda')),
            ('BACKGROUND', (1, 1), (1, 1), colors.HexColor('#f8d7da')),
            ('BACKGROUND', (2, 1), (2, 1),
             colors.HexColor('#d1ecf1') if saldo >= 0 else colors.HexColor('#f8d7da')),
            ('TEXTCOLOR',  (0, 1), (-1, 1), colors.HexColor('#333333')),
            ('FONTNAME',   (0, 1), (-1, 1), 'Helvetica-Bold'),
            ('FONTSIZE',   (0, 1), (-1, 1), 14),
            ('GRID',       (0, 0), (-1, -1), 1, colors.grey),
            ('TOPPADDING', (0, 1), (-1, 1), 10),
            ('BOTTOMPADDING', (0, 1), (-1, 1), 10),
        ]))
        elements.append(summary_table)
        elements.append(Spacer(1, 0.8 * cm))

        sorted_transactions = sorted(transactions, key=lambda x: x['date'], reverse=True)
        table_data = [['Data', 'Tipo', 'Categoria', 'Fornitore', 'Importo']]
        for trans in sorted_transactions:
            table_data.append([
                trans['date'],
                trans['type'],
                trans.get('service', 'N/A'),
                trans.get('provider', 'N/A'),
                f"€ {trans['amount']:,.2f}"
            ])

        transactions_table = Table(
            table_data,
            colWidths=[3 * cm, 3 * cm, 5 * cm, 6 * cm, 3.5 * cm]
        )
        table_style = [
            ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#2c3e50')),
            ('TEXTCOLOR',  (0, 0), (-1, 0), colors.whitesmoke),
            ('ALIGN',      (0, 0), (-1, -1), 'LEFT'),
            ('ALIGN',      (4, 0), (4, -1), 'RIGHT'),
            ('FONTNAME',   (0, 0), (-1, 0), 'Helvetica-Bold'),
            ('FONTSIZE',   (0, 0), (-1, 0), 10),
            ('BOTTOMPADDING', (0, 0), (-1, 0), 10),
            ('TOPPADDING', (0, 0), (-1, 0), 10),
            ('FONTNAME',   (0, 1), (-1, -1), 'Helvetica'),
            ('FONTSIZE',   (0, 1), (-1, -1), 9),
            ('GRID',       (0, 0), (-1, -1), 0.5, colors.grey),
            ('ROWBACKGROUNDS', (0, 1), (-1, -1), [colors.white, colors.HexColor('#f8f9fa')]),
            ('TOPPADDING', (0, 1), (-1, -1), 6),
            ('BOTTOMPADDING', (0, 1), (-1, -1), 6),
        ]
        for i, trans in enumerate(sorted_transactions, start=1):
            if trans['type'] == 'Entrata':
                table_style.append(('TEXTCOLOR', (4, i), (4, i), colors.HexColor('#2ecc71')))
            else:
                table_style.append(('TEXTCOLOR', (4, i), (4, i), colors.HexColor(COLORE_ERROR)))

        transactions_table.setStyle(TableStyle(table_style))
        elements.append(transactions_table)

        doc.build(elements)
        return str(filepath)

    def export_to_excel(self, transactions, property_name=None, start_date=None, end_date=None):
        """Esporta transazioni in Excel con formattazione"""
        timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
        filepath  = self.exports_dir / f"transazioni_{timestamp}.xlsx"

        wb = Workbook()
        wb.remove(wb.active)

        # ── Foglio 1: Riepilogo ──────────────────────────────────
        ws_summary  = wb.create_sheet("Riepilogo")
        title_font  = Font(name='Arial', size=16, bold=True, color='1E7BE7')
        header_font = Font(name='Arial', size=12, bold=True, color='FFFFFF')
        header_fill = PatternFill(start_color='2C3E50', end_color='2C3E50', fill_type='solid')
        border      = Border(
            left  =Side(style='thin'), right =Side(style='thin'),
            top   =Side(style='thin'), bottom=Side(style='thin')
        )

        ws_summary['A1'] = '📊 Report Transazioni'
        ws_summary['A1'].font = title_font
        ws_summary.merge_cells('A1:E1')

        ws_summary['A3'] = f"Generato il: {datetime.now().strftime('%d/%m/%Y %H:%M')}"
        if property_name:
            ws_summary['A4'] = f"Proprietà: {property_name}"
        if start_date and end_date:
            ws_summary['A5'] = f"Periodo: {start_date} - {end_date}"

        totale_entrate = sum(t['amount'] for t in transactions if t['type'] == 'Entrata')
        totale_uscite  = sum(t['amount'] for t in transactions if t['type'] == 'Uscita')
        saldo          = totale_entrate - totale_uscite

        for row, label, value, color in [
            (8, 'Totale Entrate', totale_entrate, 'D4EDDA'),
            (9, 'Totale Uscite',  totale_uscite,  'F8D7DA'),
            (10, 'Saldo Netto',   saldo,          'D1ECF1' if saldo >= 0 else 'F8D7DA'),
        ]:
            if row == 8:
                for col in ['A', 'B']:
                    cell = ws_summary[f'{col}7']
                    cell.value = 'Tipo' if col == 'A' else 'Importo'
                    cell.font  = header_font
                    cell.fill  = header_fill
                    cell.alignment = Alignment(horizontal='center')
                    cell.border    = border

            ws_summary[f'A{row}'] = label
            ws_summary[f'B{row}'] = value
            ws_summary[f'B{row}'].number_format = '€#,##0.00'
            ws_summary[f'B{row}'].fill = PatternFill(
                start_color=color, end_color=color, fill_type='solid'
            )
            if label == 'Saldo Netto':
                ws_summary[f'B{row}'].font = Font(bold=True)
            for col in ['A', 'B']:
                ws_summary[f'{col}{row}'].border = border

        ws_summary.column_dimensions['A'].width = 20
        ws_summary.column_dimensions['B'].width = 18

        # ── Foglio 2: Transazioni ────────────────────────────────
        ws_trans = wb.create_sheet("Transazioni")
        headers  = ['Data', 'Tipo', 'Categoria', 'Fornitore', 'Importo']
        for col, header in enumerate(headers, start=1):
            cell           = ws_trans.cell(row=1, column=col, value=header)
            cell.font      = header_font
            cell.fill      = header_fill
            cell.alignment = Alignment(horizontal='center')
            cell.border    = border

        sorted_transactions = sorted(transactions, key=lambda x: x['date'], reverse=True)
        for row_idx, trans in enumerate(sorted_transactions, start=2):
            ws_trans.cell(row=row_idx, column=1, value=trans['date'])

            tipo_cell = ws_trans.cell(row=row_idx, column=2, value=trans['type'])
            color     = '2ECC71' if trans['type'] == 'Entrata' else 'E74C3C'
            tipo_cell.font = Font(color=color, bold=True)

            ws_trans.cell(row=row_idx, column=3, value=trans.get('service', 'N/A'))
            ws_trans.cell(row=row_idx, column=4, value=trans.get('provider', 'N/A'))

            amount_cell               = ws_trans.cell(row=row_idx, column=5, value=trans['amount'])
            amount_cell.number_format = '€#,##0.00'
            amount_cell.font          = Font(color=color, bold=True)

            bg = 'FFFFFF' if row_idx % 2 == 0 else 'F8F9FA'
            for col in range(1, 6):
                cell           = ws_trans.cell(row=row_idx, column=col)
                cell.border    = border
                cell.fill      = PatternFill(start_color=bg, end_color=bg, fill_type='solid')
                cell.alignment = Alignment(horizontal='left' if col < 5 else 'right')

        for col, width in zip(['A', 'B', 'C', 'D', 'E'], [12, 12, 25, 30, 15]):
            ws_trans.column_dimensions[col].width = width

        wb.save(str(filepath))
        return str(filepath)