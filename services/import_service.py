"""
Servizio per importazione transazioni da Excel
Supporta formati .xlsx e .csv con validazione completa
"""
import os
from datetime import datetime
from pathlib import Path
import openpyxl
import csv

from validation_utils import (
    parse_decimal, 
    validate_required_text, 
    validate_transaction_type,
    validate_date_string,
    ValidationError
)
from security_manager import SecurityManager


class ImportService:
    """Gestisce l'importazione di transazioni da file Excel/CSV"""
    
    def __init__(self, transaction_service, property_service, supplier_service, logger):
        self.transaction_service = transaction_service
        self.property_service = property_service
        self.supplier_service = supplier_service
        self.logger = logger
        self.security = SecurityManager()
    
    def import_from_excel(self, file_path, property_id):
        """
        Importa transazioni da file Excel
        
        Args:
            file_path: Path del file Excel
            property_id: ID proprietà a cui associare le transazioni
            
        Returns:
            dict: {
                'success': int,
                'failed': int,
                'errors': list,
                'imported_ids': list
            }
        """
        result = {
            'success': 0,
            'failed': 0,
            'errors': [],
            'imported_ids': []
        }
        
        # Valida file
        validation = self.security.validate_file_upload(
            file_path,
            allowed_extensions={'xlsx', 'xls', 'csv'}
        )
        
        if not validation['valid']:
            result['errors'].append(f"File non valido: {validation['error']}")
            return result
        
        # Determina tipo file
        extension = validation['extension']
        
        try:
            if extension in ['xlsx', 'xls']:
                transactions = self._parse_excel(file_path)
            elif extension == 'csv':
                transactions = self._parse_csv(file_path)
            else:
                result['errors'].append(f"Estensione non supportata: {extension}")
                return result
            
            # Importa transazioni
            for row_num, trans_data in enumerate(transactions, start=2):
                try:
                    # Valida e importa
                    validated_data = self._validate_transaction_data(
                        trans_data, 
                        property_id,
                        row_num
                    )
                    
                    # Crea transazione
                    trans_id = self.transaction_service.create_with_supplier(
                        property_id=validated_data['property_id'],
                        date=validated_data['date'],
                        trans_type=validated_data['type'],
                        amount=validated_data['amount'],
                        provider=validated_data['provider'],
                        service=validated_data['service'],
                        supplier_id=validated_data.get('supplier_id')
                    )
                    
                    if trans_id:
                        result['success'] += 1
                        result['imported_ids'].append(trans_id)
                        self.logger.info(f"Transazione importata: {trans_id} (riga {row_num})")
                    else:
                        result['failed'] += 1
                        result['errors'].append(f"Riga {row_num}: Impossibile salvare nel database")
                        
                except ValidationError as e:
                    result['failed'] += 1
                    result['errors'].append(f"Riga {row_num}: {str(e)}")
                    self.logger.warning(f"Errore validazione riga {row_num}: {e}")
                except Exception as e:
                    result['failed'] += 1
                    result['errors'].append(f"Riga {row_num}: Errore imprevisto - {str(e)}")
                    self.logger.error(f"Errore importazione riga {row_num}: {e}")
            
            return result
            
        except Exception as e:
            self.logger.error(f"Errore lettura file: {e}")
            result['errors'].append(f"Errore lettura file: {str(e)}")
            return result
    
    def _parse_excel(self, file_path):
        """
        Estrae transazioni da file Excel
        
        Formato atteso:
        - Colonna A: Data (dd/MM/yyyy)
        - Colonna B: Tipo (Entrata/Uscita)
        - Colonna C: Categoria/Servizio
        - Colonna D: Fornitore
        - Colonna E: Importo (numero)
        - Colonna F (opzionale): Nome Fornitore esistente
        
        Returns:
            list: Lista di dict con dati transazioni
        """
        transactions = []
        
        try:
            workbook = openpyxl.load_workbook(file_path, data_only=True)
            sheet = workbook.active
            
            # Salta header (prima riga)
            for row in sheet.iter_rows(min_row=2, values_only=True):
                # Salta righe vuote
                if not any(row):
                    continue
                
                # Estrai dati (gestisci celle vuote)
                data = {
                    'date': row[0] if len(row) > 0 else None,
                    'type': row[1] if len(row) > 1 else None,
                    'service': row[2] if len(row) > 2 else None,
                    'provider': row[3] if len(row) > 3 else None,
                    'amount': row[4] if len(row) > 4 else None,
                    'supplier_name': row[5] if len(row) > 5 else None
                }
                
                transactions.append(data)
            
            workbook.close()
            
        except Exception as e:
            self.logger.error(f"Errore parsing Excel: {e}")
            raise ValueError(f"Impossibile leggere il file Excel: {str(e)}")
        
        return transactions
    
    def _parse_csv(self, file_path):
        """
        Estrae transazioni da file CSV
        
        Stesso formato dell'Excel ma in CSV
        """
        transactions = []
        
        try:
            with open(file_path, 'r', encoding='utf-8-sig') as csvfile:
                # Prova a rilevare il delimitatore
                sample = csvfile.read(1024)
                csvfile.seek(0)
                
                try:
                    dialect = csv.Sniffer().sniff(sample)
                    reader = csv.reader(csvfile, dialect)
                except:
                    reader = csv.reader(csvfile, delimiter=',')
                
                # Salta header
                next(reader, None)
                
                for row in reader:
                    # Salta righe vuote
                    if not any(row):
                        continue
                    
                    data = {
                        'date': row[0] if len(row) > 0 else None,
                        'type': row[1] if len(row) > 1 else None,
                        'service': row[2] if len(row) > 2 else None,
                        'provider': row[3] if len(row) > 3 else None,
                        'amount': row[4] if len(row) > 4 else None,
                        'supplier_name': row[5] if len(row) > 5 else None
                    }
                    
                    transactions.append(data)
        
        except Exception as e:
            self.logger.error(f"Errore parsing CSV: {e}")
            raise ValueError(f"Impossibile leggere il file CSV: {str(e)}")
        
        return transactions
    
    def _validate_transaction_data(self, data, property_id, row_num):
        """
        Valida i dati di una transazione
        
        Args:
            data: Dict con dati grezzi
            property_id: ID proprietà
            row_num: Numero riga (per messaggi errore)
            
        Returns:
            dict: Dati validati
            
        Raises:
            ValidationError: Se validazione fallisce
        """
        validated = {}
        
        # Property ID
        validated['property_id'] = property_id
        
        # Data
        if not data.get('date'):
            raise ValidationError("Data mancante")
        
        # Gestisci diversi formati data
        date_value = data['date']
        
        if isinstance(date_value, datetime):
            # Se è già un datetime (da Excel)
            validated['date'] = date_value.strftime('%d/%m/%Y')
        elif isinstance(date_value, str):
            # Se è stringa, valida formato
            validated['date'] = validate_date_string(date_value, "Data")
        else:
            raise ValidationError(f"Formato data non riconosciuto: {type(date_value)}")
        
        # Tipo
        if not data.get('type'):
            raise ValidationError("Tipo mancante")
        
        type_value = str(data['type']).strip()
        validated['type'] = validate_transaction_type(type_value)
        
        # Servizio/Categoria
        if not data.get('service'):
            raise ValidationError("Categoria/Servizio mancante")
        
        validated['service'] = validate_required_text(
            str(data['service']),
            "Servizio",
            min_length=2,
            max_length=200
        )
        
        # Fornitore
        if not data.get('provider'):
            raise ValidationError("Fornitore mancante")
        
        validated['provider'] = validate_required_text(
            str(data['provider']),
            "Fornitore",
            min_length=2,
            max_length=200
        )
        
        # Importo
        if data.get('amount') is None or data.get('amount') == '':
            raise ValidationError("Importo mancante")
        
        # Gestisci importo come numero o stringa
        amount_value = data['amount']
        
        if isinstance(amount_value, (int, float)):
            validated['amount'] = float(amount_value)
        elif isinstance(amount_value, str):
            validated['amount'] = parse_decimal(amount_value, "Importo")
        else:
            raise ValidationError(f"Formato importo non riconosciuto: {type(amount_value)}")
        
        # Verifica range
        if validated['amount'] <= 0:
            raise ValidationError("Importo deve essere maggiore di zero")
        
        # Fornitore esistente (opzionale)
        validated['supplier_id'] = None
        
        if data.get('supplier_name'):
            supplier_name = str(data['supplier_name']).strip()
            
            # Cerca fornitore per nome
            suppliers = self.supplier_service.search(
                supplier_name,
                category=validated['service'],
                property_id=property_id
            )
            
            if suppliers:
                # Usa il primo match
                validated['supplier_id'] = suppliers[0]['id']
                self.logger.info(f"Fornitore collegato: {suppliers[0]['name']} (ID: {suppliers[0]['id']})")
        
        return validated

    def generate_template(self, output_path=None):
        """
        Genera file Excel template per l'importazione

        Args:
            output_path: Path dove salvare il template (opzionale)

        Returns:
            str: Path del file generato
        """
        if output_path is None:
            output_path = os.path.join(
                'exports',
                f'template_importazione_{datetime.now().strftime("%Y%m%d_%H%M%S")}.xlsx'
            )

        dir_name = os.path.dirname(output_path)
        if dir_name:
            os.makedirs(dir_name, exist_ok=True)

        try:
            workbook = openpyxl.Workbook()
            sheet = workbook.active
            sheet.title = "Transazioni"

            headers = [
                'Data (dd/MM/yyyy)*',
                'Tipo (Entrata/Uscita)*',
                'Categoria/Servizio*',
                'Fornitore*',
                'Importo*',
                'Nome Fornitore Esistente'
            ]

            for col, header in enumerate(headers, start=1):
                cell = sheet.cell(row=1, column=col, value=header)
                cell.font = openpyxl.styles.Font(bold=True)
                cell.fill = openpyxl.styles.PatternFill(
                    start_color="2C3E50",
                    end_color="2C3E50",
                    fill_type="solid"
                )
                cell.font = openpyxl.styles.Font(color="FFFFFF", bold=True)

            examples = [
                ['15/01/2024', 'Uscita', 'Bolletta Luce', 'ENEL Energia', 150.50, ''],
                ['20/01/2024', 'Entrata', 'Affitto', 'Inquilino Rossi', 800.00, ''],
                ['25/01/2024', 'Uscita', 'Manutenzione', 'Idraulico Mario', 85.00, 'Idraulico Mario']
            ]

            for row_idx, example in enumerate(examples, start=2):
                for col_idx, value in enumerate(example, start=1):
                    sheet.cell(row=row_idx, column=col_idx, value=value)

            sheet.column_dimensions['A'].width = 18
            sheet.column_dimensions['B'].width = 18
            sheet.column_dimensions['C'].width = 25
            sheet.column_dimensions['D'].width = 25
            sheet.column_dimensions['E'].width = 15
            sheet.column_dimensions['F'].width = 30

            instructions_sheet = workbook.create_sheet("Istruzioni")

            instructions = [
                ["ISTRUZIONI PER L'IMPORTAZIONE TRANSAZIONI", ""],
                ["", ""],
                ["Colonne obbligatorie (*):", ""],
                ["1. Data", "Formato: dd/MM/yyyy (es: 15/01/2024)"],
                ["2. Tipo", "Deve essere esattamente 'Entrata' o 'Uscita'"],
                ["3. Categoria/Servizio", "Nome del servizio (es: Bolletta Luce, Affitto)"],
                ["4. Fornitore", "Nome del fornitore/emittente"],
                ["5. Importo", "Numero decimale (usa punto o virgola: 150.50 o 150,50)"],
                ["", ""],
                ["Colonna opzionale:", ""],
                ["6. Nome Fornitore Esistente", "Se vuoi collegare a un fornitore già presente nel sistema"],
                ["", ""],
                ["Note importanti:", ""],
                ["- La prima riga (header) verrà ignorata", ""],
                ["- Righe vuote verranno ignorate", ""],
                ["- Puoi usare sia Excel (.xlsx) che CSV", ""],
                ["- Gli importi devono essere positivi", ""],
                ["- Le date devono essere nel formato dd/MM/yyyy", ""],
                ["", ""],
                ["In caso di errori:", ""],
                ["- Riceverai un report dettagliato con le righe che hanno causato problemi", ""],
                ["- Le transazioni valide verranno comunque importate", ""],
            ]

            for row_idx, (label, value) in enumerate(instructions, start=1):
                cell_label = instructions_sheet.cell(row=row_idx, column=1, value=label)
                instructions_sheet.cell(row=row_idx, column=2, value=value)

                if row_idx == 1:
                    cell_label.font = openpyxl.styles.Font(bold=True, size=14)
                elif any(kw in label for kw in ("obbligatorie", "opzionale", "importanti", "errori")):
                    cell_label.font = openpyxl.styles.Font(bold=True, size=12)

            instructions_sheet.column_dimensions['A'].width = 35
            instructions_sheet.column_dimensions['B'].width = 60

            workbook.save(output_path)
            workbook.close()

            self.logger.info(f"Template generato: {output_path}")
            return output_path

        except Exception as e:
            self.logger.error(f"Errore generazione template: {e}")
            raise ValueError(f"Impossibile generare template: {str(e)}")
